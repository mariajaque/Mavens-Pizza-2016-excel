# Qué hay que comprar para la semana que viene
"""
- The order_details tables has 48621 rows containing order details regarding
pizza type and order quantity.
- The orders table record the datetime indicators of the 21351 orders.
- The pizza_types table specifies the category, ingredients information about
the 33 different pizza types offered by the pizza place.
- The pizzas table has 97 rows containing the pricing details of pizza based on
the size and pizza type
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
import openpyxl
import warnings
import signal
import sys


warnings.filterwarnings("ignore")

def handler_signal(signal,frame):

    # Salida controlada del programa en caso de pulsar 
    # control C

    print("\n\n [!] out .......\ n")

    sys.exit(1)

signal.signal(signal.SIGINT,handler_signal)


def extract_csv():

    fechas = pd.read_csv('orders.csv', sep=';')
    pedidos = pd.read_csv('pizzas.csv', sep=',')
    detalles = pd.read_csv('order_details.csv', sep=';')
    ingredientes = pd.read_csv('pizza_types.csv', sep=',', encoding='Windows-1252')

    informe = informe_de_datos(fechas, pedidos, detalles, ingredientes)

    return (fechas, pedidos, detalles, ingredientes, informe)


def transform_csv(fechas, pedidos, detalles, ingredientes):

    # Primero ordenamos los id y posteriormente reestablecemos
    # los índices del dataframe

    fechas = fechas.sort_values('order_id')
    fechas.index = [i for i in range(fechas.shape[0])]

    detalles = detalles.sort_values('order_details_id')
    detalles = detalles.dropna()
    detalles.index = [i for i in range(detalles.shape[0])]

    # En fechas no transformamos las horas ya que esa columna no se va a
    # utilizar para nada
    # Transformamos las fechas todas al mismo formato
    # Si en alguno da error lo cambiamos por un null

    for i in fechas.index:

        fechas.loc[i, 'date'] = pd.to_datetime(fechas['date'].iloc[i], errors='coerce')

    # Vamos a sustituir los nats por el valor que se haya podido transformar
    # antes

    fi = pd.to_datetime('2016-01-01')

    for i in fechas.index:

        if str(fechas['date'].iloc[i]) == str(pd.NaT):
            fechas.loc[i, 'date'] = fi

        else:
            fi = fechas['date'].iloc[i]

    # Para el dataframe de detalles, primero nos quitaremos todos los NaN
    # y posteriormente nos quitaremos todos los negativos en la columna de
    # orders, reemplazándolos por su valor absoluto -> Asumimos que se
    # equivocaron al introducir los datos

    # detalles = detalles.dropna()

    # Reemplazamos os números escritos con letras por números enteros
    # Habiendo visto los datos los únicos números que aparecen a mano
    # son one y two

    detalles['quantity'].replace(to_replace=r'[O-o][N-n][E-e]', value=1, regex=True,inplace=True)
    detalles['quantity'].replace(to_replace=r'[T-t][W-w][O-o]', value=2, regex=True,inplace=True)

    # Obtengo los índices de aquellos números negativos en cantidad

    for i in detalles.index:

        try:
            detalles.loc[i, 'quantity'] = abs(int(detalles['quantity'].iloc[i]))
        except:
            ...

    detalles['pizza_id'] = detalles['pizza_id'].str.replace(' ', '_')
    detalles['pizza_id'] = detalles['pizza_id'].str.replace('-', '_')
    detalles['pizza_id'] = detalles['pizza_id'].str.replace('@', 'a')
    detalles['pizza_id'] = detalles['pizza_id'].str.replace('0', 'o')
    detalles['pizza_id'] = detalles['pizza_id'].str.replace('3', 'e')

    # Vamos a querer tener todos los datos en un único dataframe
    # Modificaremos el de order details pues es el más completo
    # Le añadiremos una nueva columna que sea el número de la semana
    # asociado a la fecha del pedido. Añadiremos una
    # columna por cada posible ingrediente de la pizza

    dias = []
    num_semana = []

    for fecha in fechas['date']:
        dia = pd.to_datetime(fecha, dayfirst=True)
        dias.append(dia.day_of_week)
        num_semana.append(dia.week)

    fechas['semana'] = num_semana
    fechas['dia_semana'] = dias

    # Nos guardamos para cada order_id en detalles su fecha
    # asociada

    semanas = []
    dia_semana = []

    for s in detalles['order_id']:

        indice = fechas[fechas['order_id'] == s].index
        semana = fechas['semana'].iloc[indice]
        d = fechas['dia_semana'].iloc[indice]

        semanas.append(int(semana))
        dia_semana.append(int(d))

    detalles['semana'] = semanas
    detalles['dia'] = dia_semana

    # Obtenemos todos los posibles ingredientes que emplea
    # en la elaboración de sus pizzas

    lista_ingredientes = []
    for ingrediente in ingredientes['ingredients']:
        varios = ingrediente.split(',')
        lista_ingredientes += varios

    set_ingredientes = set(lista_ingredientes)

    # Creamos una columna por cada ingrediente en detalles
    # Almacenamos en un diccionario el índice de cada ingrediente

    indices = dict()

    for i in set_ingredientes:
        detalles[i] = [0 for i in detalles.index]
        indice = detalles.columns.get_loc(i)
        indices[i] = indice

    # Para cada tipo de pizza en order detail, les sumamos
    # las cantidades a sus ingredientes correspondientes
    # Para las s sumaremos una unidad de cada ingrediente
    # Para las m sumaremos 2 y para las L sumaremos 3

    tipos_de_pizzas = pedidos['pizza_id'].tolist()
    tamanos = ['s', 'm', 'l', 'xl', 'xxl']
    ing_asociados = dict()

    for tipo in tipos_de_pizzas:

        tamano = tipo.split('_')[-1]
        ingredientes_str = ingredientes[ingredientes['pizza_type_id'] == tipo[:-len(tamano)-1]]['ingredients'].tolist()[0]
        lista_ingredientes_comprar = ingredientes_str.split(',')
        ing_asociados[tipo] = lista_ingredientes_comprar

    # Sumamos la cantidad de cada ingrediente que ha necesitado cada pedido

    for i in detalles.index:

        try:
            pedido = detalles['pizza_id'].iloc[i]
            cantidad = detalles['quantity'].iloc[i]
            ing = ing_asociados[pedido]
            tamano = pedido.split('_')[-1]

            for j in ing:
                detalles.loc[i, j] += cantidad * (tamanos.index(tamano) + 1)
        
        except:
            ...

    return detalles


def load_csv(datos):

    # El dataframe obtenido en el transform csv
    pass


def extract():

    # Extrae los datos finales ya trabajados de la pizería
    pass


def transform(datos):

    ingredientes = datos.columns.values
    ingredientes = ingredientes[6:]

    # Nuestro predict será la media de las modas de cada ingrediente

    suma_semana = datos.pivot_table(index='semana', aggfunc='sum')
    suma_semana_ingredientes = suma_semana[ingredientes]
    modas = suma_semana_ingredientes.mode().mean().round().tolist()

    # Creamos un dataframe con el valor calculado para cada ingrediente

    d = {'Ingredientes:': ingredientes, 'Unidades a comprar:': modas}
    res = pd.DataFrame(data=d)

    return res


def load(res, datos):

    anadir_datos(res, datos)
    hoja_reporte_ejecutivo()
    hoja_reporte_ingredientes()
    hoja_reportes_pedidos()


def informe_de_datos(fechas, pedidos, detalles, ingredientes):

    # Primero vemos el número de NaNs y de Nulls de cada df
    # Agregamos también el tipo de cada columna

    fichero = open('informe_calidad_datos.txt', 'w')
    informe = {}

    dfs = [fechas, pedidos, detalles, ingredientes]
    nombres = ['orders.csv', 'pizzas.csv', 'order_details.csv', 'pizza_types.csv']

    for df in range(len(dfs)):

        valores = {}

        null = {}
        nan = {}

        columnas = dfs[df].columns.values.tolist()

        tipos_columna = {}

        for c in columnas:

            tipos = dfs[df][c].dtypes
            nulls = dfs[df][c].isnull().sum()
            nans = dfs[df][c].isna().sum()

            tipos_columna[c] = tipos
            null[c] = nulls
            nan[c] = nans

        valores['Nulls'] = null
        valores['NaNs'] = nan
        valores['Tipos'] = tipos_columna

        informe[nombres[df]] = valores

    return informe


def anadir_datos(res, datos):

    # Esta función añade las tablas de datos que queramos a las distintas hojas del fichero
    # Reporte ingredientes:

    ingredientes = res['Ingredientes:'].values.tolist()

    # Creamos un df con el total de cada ingrediente
    ing = pd.DataFrame({'quantity':datos[ingredientes].sum()})

    tabla_ingredientes = ing.pivot_table(index=ing.index, aggfunc='sum')
    tabla_ingredientes = tabla_ingredientes.sort_values('quantity', ascending=False)

    # Reporte pedidos:

    tabla_pizzas = datos.pivot_table(index='pizza_id', values='quantity', aggfunc='sum')
    tabla_pizzas = tabla_pizzas.sort_values('quantity', ascending=False)

    # reporte ejecutivo:

    tabla_comprar = res.pivot_table(index='Ingredientes:', values='Unidades a comprar:')


    # Añadimos las tablas al excel

    with pd.ExcelWriter('mavens_pizza.xlsx') as escribir:

        tabla_comprar.to_excel(escribir, startrow=4, startcol=1, sheet_name='Reporte Ejecutivo')
        tabla_pizzas.to_excel(escribir, startrow=4, startcol=1, sheet_name='Reporte de Pedidos')
        tabla_ingredientes.to_excel(escribir, startrow=4, startcol=1, sheet_name='Reporte de Ingredientes')


def hoja_reporte_ejecutivo():

    # Guardo el valor de la tabla y el contenido del fichero

    fichero_excel = load_workbook('mavens_pizza.xlsx')
    reporte_ingredientes = fichero_excel['Reporte Ejecutivo']

    # Se obtienen los límites en los que se encuentra la tabla de información

    columna_min = reporte_ingredientes.min_column
    columna_max = reporte_ingredientes.max_column

    fila_min = reporte_ingredientes.min_row
    fila_max = reporte_ingredientes.max_row

    # Creo un gráfico de barras con las cantidades que se deben comprar de cada ingrediente

    grafico_barras_ing = BarChart()
    data = Reference(reporte_ingredientes, min_col=columna_min + 1, max_col=columna_max, min_row=fila_min, max_row=fila_max)
    nombre_ing = Reference(reporte_ingredientes, min_col=columna_min, max_col=columna_min, min_row=fila_min + 1, max_row=fila_max)

    grafico_barras_ing.add_data(data, titles_from_data=True)
    grafico_barras_ing.set_categories(nombre_ing)

    # Añadimos el barchart al arcivo excel
    # Le indicamos que esté en la celda F5

    reporte_ingredientes.add_chart(grafico_barras_ing, 'F5')

    # Le metemos también el estilo de la gráfica y el título
    # El estilo 2 nos pone la gráfica en azul

    grafico_barras_ing.title = 'Ingredientes a comprar'
    grafico_barras_ing.style = 2

    # Agregamos los títulos de la hoja de excel

    reporte_ingredientes['B2'] = 'Reporte Ejecutivo'
    reporte_ingredientes['B2'].font = Font('Calibri', bold=True, size = 20)

    reporte_ingredientes['B3'] = '2016'
    reporte_ingredientes['B3'].font = Font('Calibri', bold=True, size = 12)

    # Guardamos los cambios
    fichero_excel.save('mavens_pizza.xlsx')


def hoja_reporte_ingredientes():

    fichero_excel = load_workbook('mavens_pizza.xlsx')

    # Para conseguir estar en la hoja que habíamos creado

    reporte_ingredientes = fichero_excel['Reporte de Ingredientes']

    # Sacamos las columnas mínima y máxima

    columna_min = reporte_ingredientes.min_column
    columna_max = reporte_ingredientes.max_column

    # Sacamos las filas mínima y máxima

    fila_min = reporte_ingredientes.min_row
    fila_max = reporte_ingredientes.max_row


    # Para añadir una gráfica empleamos BarChart y Reference de la librería openpyxl.chart
    # Sumamos 1 a columna minima ya que los datos que queremos se encuentras en la columna C
    # y en la columna B solo encontramos los nombres de las pizzas
    # En la parte de pizza_id sumamos 1 a la mínima fila ya que en la mínima fila solo se encuentra
    # el nombre de la variable cantidad pero no hay datos. También le ponemos que busque las categorías
    # únicamente en la primera columna ya que es donde están

    grafico_barras_ing = BarChart()
    data = Reference(reporte_ingredientes, min_col=columna_min + 1, max_col=columna_max, min_row=fila_min, max_row=fila_max)
    nombre_ing = Reference(reporte_ingredientes, min_col=columna_min, max_col=columna_min, min_row=fila_min + 1, max_row=fila_max)

    # Agregamos al gráfico de barras que ya teníamos creado los datos que acabamos de extraer
    # Dado que los títulos se encuentran en la parte de datos lo agregamos como true

    grafico_barras_ing.add_data(data, titles_from_data=True)
    grafico_barras_ing.set_categories(nombre_ing)

    # Añadimos el barchart al arcivo excel
    # Le indicamos que esté en la celda F5

    reporte_ingredientes.add_chart(grafico_barras_ing, 'F5')

    # Le metemos también el estilo de la gráfica y el título
    # El estilo 2 nos pone la gráfica en azul

    grafico_barras_ing.title = 'Total de cada tipo de cada ingrediente pedido en un año'
    grafico_barras_ing.style = 2

    # Buscamos las 5 celdas con valores máximos
    # Como los datos están ordenados serán las 5 primeras celdas (filas de la 6 a la 10)
    # Las pintamos de verde

    for i in range(6, 11):
        reporte_ingredientes['C' + str(i)].fill = PatternFill(fill_type='solid', fgColor='58FA58')
        reporte_ingredientes['B' + str(i)].fill = PatternFill(fill_type='solid', fgColor='01DF01')

    # Hacemos lo mismo con las 5 pizzas menos pedidas
    # Para ello tomamos las 5 últimas filas (de la 92 a la 96)
    # Las pintamos de rojo

    for i in range(fila_max - 4, fila_max + 1):
        reporte_ingredientes['C' + str(i)].fill = PatternFill(fill_type='solid', fgColor='FA5858')
        reporte_ingredientes['B' + str(i)].fill = PatternFill(fill_type='solid', fgColor='DF0101')


    # Añadimos un barplot con las 5 peores pizzas y otro con las 5 mejores
    # 5 MEJORES PIZZAS:

    grafico_barras_ing_top = BarChart()
    data = Reference(reporte_ingredientes, min_col=columna_min + 1, max_col=columna_max, min_row=5, max_row=10)
    nombre_ing = Reference(reporte_ingredientes, min_col=columna_min, max_col=columna_min, min_row=6, max_row=10)

    grafico_barras_ing_top.add_data(data, titles_from_data=True)
    grafico_barras_ing_top.set_categories(nombre_ing)

    reporte_ingredientes.add_chart(grafico_barras_ing_top, 'F21')

    grafico_barras_ing_top.title = 'Top 5 ingredientes'
    grafico_barras_ing_top.style = 5


    # 5 PEORES PIZZAS:

    grafico_barras_ing_worst = BarChart()
    data = Reference(reporte_ingredientes, min_col=columna_min + 1, max_col=columna_max, min_row=fila_max - 4, max_row=fila_max) 
    nombre_ing = Reference(reporte_ingredientes, min_col=columna_min, max_col=columna_min, min_row=fila_max - 4, max_row=fila_max)

    grafico_barras_ing_worst.add_data(data)
    grafico_barras_ing_worst.set_categories(nombre_ing)

    reporte_ingredientes.add_chart(grafico_barras_ing_worst, 'F37')

    grafico_barras_ing_worst.title = '5 ingredientes menos vendidos'
    grafico_barras_ing_worst.style = 4

    # Nos sacamos el total de pizzas que se han comprado

    reporte_ingredientes['C' + str(fila_max + 1)] = f'=SUM(C3:C{fila_max})'
    reporte_ingredientes['C' + str(fila_max + 1)].fill = PatternFill(fill_type='solid', fgColor='F3F781') # Color amarillo claro
    reporte_ingredientes['B' + str(fila_max + 1)] = 'Total de pedidos'
    reporte_ingredientes['B' + str(fila_max + 1)].fill = PatternFill(fill_type='solid', fgColor='FFFF00') # color amarillo


    # Agregamos los títulos de la hoja de excel

    reporte_ingredientes['B2'] = 'Reporte de Ingredientes'
    reporte_ingredientes['B2'].font = Font('Calibri', bold=True, size = 20)

    reporte_ingredientes['B3'] = '2016'
    reporte_ingredientes['B3'].font = Font('Calibri', bold=True, size = 12)


    # Guardamos el gráfico en el excel
    fichero_excel.save('mavens_pizza.xlsx')


def hoja_reportes_pedidos():

    fichero_excel = load_workbook('mavens_pizza.xlsx')
    reporte_pedidos = fichero_excel['Reporte de Pedidos']

    # Sacamos las columnas mínima y máxima

    columna_min = reporte_pedidos.min_column
    columna_max = reporte_pedidos.max_column

    # Sacamos las filas mínima y máxima

    fila_min = reporte_pedidos.min_row
    fila_max = reporte_pedidos.max_row

    # Para añadir una gráfica empleamos BarChart y Reference de la librería openpyxl.chart
    # Sumamos 1 a columna minima ya que los datos que queremos se encuentras en la columna C
    # y en la columna B solo encontramos los nombres de las pizzas
    # En la parte de pizza_id sumamos 1 a la mínima fila ya que en la mínima fila solo se encuentra
    # el nombre de la variable cantidad pero no hay datos. También le ponemos que busque las categorías
    # únicamente en la primera columna ya que es donde están

    grafico_barras_pizza = BarChart()
    data = Reference(reporte_pedidos, min_col=columna_min + 1, max_col=columna_max, min_row=fila_min, max_row=fila_max)
    pizzas_id = Reference(reporte_pedidos, min_col=columna_min, max_col=columna_min, min_row=fila_min + 1, max_row=fila_max)

    # Agregamos al gráfico de barras que ya teníamos creado los datos que acabamos de extraer
    # Dado que los títulos se encuentran en la parte de datos lo agregamos como true

    grafico_barras_pizza.add_data(data, titles_from_data=True)
    grafico_barras_pizza.set_categories(pizzas_id)

    # Añadimos el barchart al arcivo excel
    # Le indicamos que esté en la celda F5

    reporte_pedidos.add_chart(grafico_barras_pizza, 'F5')

    # Le metemos también el estilo de la gráfica y el título
    # El estilo 2 nos pone la gráfica en azul

    grafico_barras_pizza.title = 'Total de cada tipo de pizza pedido en un año'
    grafico_barras_pizza.style = 2

    # Buscamos las 5 celdas con valores máximos
    # Como los datos están ordenados serán las 5 primeras celdas (filas de la 6 a la 10)
    # Las pintamos de verde

    for i in range(6, 11):
        reporte_pedidos['C' + str(i)].fill = PatternFill(fill_type='solid', fgColor='58FA58')
        reporte_pedidos['B' + str(i)].fill = PatternFill(fill_type='solid', fgColor='01DF01')

    # Hacemos lo mismo con las 5 pizzas menos pedidas
    # Para ello tomamos las 5 últimas filas (de la 92 a la 96)
    # Las pintamos de rojo

    for i in range(92, 97):
        reporte_pedidos['C' + str(i)].fill = PatternFill(fill_type='solid', fgColor='FA5858')
        reporte_pedidos['B' + str(i)].fill = PatternFill(fill_type='solid', fgColor='DF0101')


    # Añadimos un barplot con las 5 peores pizzas y otro con las 5 mejores
    # 5 MEJORES PIZZAS:

    grafico_barras_pizza_top = BarChart()
    data = Reference(reporte_pedidos, min_col=columna_min + 1, max_col=columna_max, min_row=5, max_row=10)
    pizzas_id = Reference(reporte_pedidos, min_col=columna_min, max_col=columna_min, min_row=6, max_row=10)

    grafico_barras_pizza_top.add_data(data, titles_from_data=True)
    grafico_barras_pizza_top.set_categories(pizzas_id)

    reporte_pedidos.add_chart(grafico_barras_pizza_top, 'F21')

    grafico_barras_pizza_top.title = 'Top 5 pizzas'
    grafico_barras_pizza_top.style = 5


    # 5 PEORES PIZZAS:

    grafico_barras_pizza_worst = BarChart()
    data = Reference(reporte_pedidos, min_col=columna_min + 1, max_col=columna_max, min_row=92, max_row=96)  # Al cambiar a 96 si que pinta las 4 que nos faltan
    pizzas_id = Reference(reporte_pedidos, min_col=columna_min, max_col=columna_min, min_row=92, max_row=96)

    grafico_barras_pizza_worst.add_data(data)
    grafico_barras_pizza_worst.set_categories(pizzas_id)

    reporte_pedidos.add_chart(grafico_barras_pizza_worst, 'F37')

    grafico_barras_pizza_worst.title = '5 Pizzas menos vendidas'
    grafico_barras_pizza_worst.style = 4

    # Nos sacamos el total de pizzas que se han comprado

    reporte_pedidos['C' + str(fila_max + 1)] = '=SUM(C3:C93)'
    reporte_pedidos['C' + str(fila_max + 1)].fill = PatternFill(fill_type='solid', fgColor='F3F781') # Color amarillo claro
    reporte_pedidos['B' + str(fila_max + 1)] = 'Total de pedidos'
    reporte_pedidos['B' + str(fila_max + 1)].fill = PatternFill(fill_type='solid', fgColor='FFFF00') # color amarillo


    # Agregamos los títulos de la hoja de excel

    reporte_pedidos['B2'] = 'Reporte de Pizzas'
    reporte_pedidos['B2'].font = Font('Calibri', bold=True, size = 20)

    reporte_pedidos['B3'] = '2016'
    reporte_pedidos['B3'].font = Font('Calibri', bold=True, size = 12)


    # Guardamos el gráfico en el excel
    fichero_excel.save('mavens_pizza.xlsx')


if __name__ == '__main__':

    fechas, pedidos, detalles, ingredientes, informe = extract_csv()
    datos = transform_csv(fechas, pedidos, detalles, ingredientes)
    res = transform(datos)
    load(res, datos)
